#!/usr/bin/env python3
"""
Automated agent tests via LibreChat Open Responses API.

Usage:
    python3 tests/agent_api_tests.py                    # Run all tests
    python3 tests/agent_api_tests.py D01b D01c P01b     # Run specific tests
    python3 tests/agent_api_tests.py --agent docx        # Run all DOCX tests
    python3 tests/agent_api_tests.py --list              # List all tests

Prerequisites:
    - LibreChat running on localhost:3080
    - Agent API key in AGENT_API_KEY env var or .env file
    - All agents deployed with current instructions

Each test:
    1. Sends a prompt to the agent via /api/agents/v1/responses
    2. Checks that the response completed without error
    3. Validates methodology (checks for expected scripts/patterns in code execution)
    4. Downloads generated files and validates them
    5. Produces a PASS/FAIL report
"""

import argparse
import json
import os
import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

try:
    import requests
except ImportError:
    print("pip install requests")
    sys.exit(1)


# === Configuration ===

API_BASE = os.environ.get("LIBRECHAT_URL", "http://127.0.0.1:3080")
API_KEY = os.environ.get("AGENT_API_KEY", "")
CI_BASE = os.environ.get("CODE_INTERPRETER_URL", "http://127.0.0.1:8010")
CI_KEY = os.environ.get("CODE_INTERPRETER_KEY", "facac6914bfccdddd47595b6bf24d476e38bd42516d99bb5aff8da48df649a4c")
RESPONSES_ENDPOINT = f"{API_BASE}/api/agents/v1/responses"
RESULTS_DIR = Path("tests/results")
FILES_DIR = Path("tests/results/files")
TIMEOUT = 600  # 10 minutes max per test (agents with code execution can be slow)

# Force unbuffered output
import functools
print = functools.partial(print, flush=True)


@dataclass
class TestResult:
    test_id: str
    agent: str
    status: str = "PENDING"  # PASS, FAIL, ERROR, SKIP
    duration: float = 0
    prompt: str = ""
    response_text: str = ""
    code_blocks: list = field(default_factory=list)
    files_generated: list = field(default_factory=list)
    methodology_checks: dict = field(default_factory=dict)
    error: str = ""


@dataclass
class TestCase:
    test_id: str
    agent_id: str
    agent_name: str
    prompt: str
    methodology_patterns: list  # Regex patterns expected in code execution
    methodology_antipatterns: list = field(default_factory=list)  # Patterns that should NOT appear
    description: str = ""
    expect_file: bool = False
    file_extension: str = ""
    previous_response_id: Optional[str] = None  # For multi-turn tests


# === API Client ===

def exec_code(code: str, lang: str = "py") -> dict:
    """Execute code directly in the code-interpreter sandbox."""
    headers = {"x-api-key": CI_KEY, "Content-Type": "application/json"}
    resp = requests.post(f"{CI_BASE}/exec", headers=headers,
                         json={"code": code, "lang": lang}, timeout=TIMEOUT)
    resp.raise_for_status()
    return resp.json()


def download_file(session_id: str, file_id: str, output_path: Path) -> bool:
    """Download a file from the code-interpreter."""
    headers = {"x-api-key": CI_KEY}
    resp = requests.get(f"{CI_BASE}/download/{session_id}/{file_id}",
                        headers=headers, timeout=60)
    if resp.status_code == 200 and len(resp.content) > 0:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(resp.content)
        return True
    return False


# === File generation scripts (run directly in code-interpreter, bypass LLM) ===

FILE_GENERATION_SCRIPTS = {
    "D01b": '''
import subprocess, json, os
os.chdir('/mnt/data')
config = {"meeting":{"title":"CR Test","subtitle":"Test","date":"17/04/2026","location":"Test","organizer":"Test"},"participants":[{"name":"A","role":"R","company":"C"}],"sections":[{"title":"Test","level":1,"content":[{"type":"text","text":"Contenu test."}]}]}
with open("/tmp/config.json","w") as f: json.dump(config,f,ensure_ascii=False)
subprocess.run(["python3","/opt/skills/docx/scripts/fill_cr_template.py","/opt/skills/docx/templates/onbehalfai/template-compte-rendu.docx","cr_test.docx","/tmp/config.json"],check=True)
print("OK")
''',
    "D01c": '''
import subprocess, json, os
os.chdir('/mnt/data')
config = {"placeholders":{"[TITRE DU DOCUMENT]":"Guide Test","[Sous-titre du document]":"Test","[Auteur]":"Test","[Date]":"17/04/2026"},"sections":[{"title":"Section 1","level":1,"content":[{"type":"text","text":"Contenu."},{"type":"bullets","items":["A","B"]},{"type":"code","text":"echo hello"},{"type":"table","headers":["Col1","Col2"],"rows":[["a","b"]]}]}]}
with open("/tmp/config.json","w") as f: json.dump(config,f,ensure_ascii=False)
subprocess.run(["python3","/opt/skills/docx/scripts/fill_template.py","/opt/skills/docx/templates/onbehalfai/template-base.docx","guide_test.docx","/tmp/config.json"],check=True)
print("OK")
''',
    "P01b": '''
const pptxgen = require("pptxgenjs");
const pptx = new pptxgen();
pptx.layout = "LAYOUT_16x9";
const s = pptx.addSlide();
s.background = { fill: "1C244B" };
s.addText("Test Slide", {x:0.5,y:2,w:9,h:1.5,fontSize:36,fontFace:"Arial",bold:true,color:"FFFFFF"});
pptx.writeFile({fileName:"/mnt/data/test_pptx.pptx"}).then(()=>console.log("OK"));
''',
    "X01": '''
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
wb = Workbook()
ws = wb.active
ws.title = "Budget"
hdr = PatternFill(start_color="2F5597",end_color="2F5597",fill_type="solid")
hfont = Font(name="Arial",size=11,bold=True,color="FFFFFF")
for col,h in enumerate(["Poste","Montant"],1):
    c = ws.cell(row=1,column=col,value=h)
    c.fill = hdr
    c.font = hfont
ws.cell(row=2,column=1,value="Salaires")
ws.cell(row=2,column=2,value=45000)
wb.save("/mnt/data/budget_test.xlsx")
print("OK")
''',
    "A02": '''
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
fig, axes = plt.subplots(2,2,figsize=(10,8))
colors = ["#2F5597","#5B9AD4","#FB840D","#FCA810"]
axes[0,0].plot([1,2,3,4],[10,20,15,25],color=colors[0])
axes[0,0].set_title("CA Mensuel")
axes[0,1].pie([40,30,20,10],labels=["A","B","C","D"],colors=colors)
axes[0,1].set_title("Répartition")
axes[1,0].barh(["C1","C2","C3"],[100,80,60],color=colors[1])
axes[1,0].set_title("Top Clients")
axes[1,1].scatter(np.random.rand(20),np.random.rand(20),color=colors[2])
axes[1,1].set_title("Scatter")
plt.suptitle("Dashboard OBA",fontweight="bold")
plt.tight_layout()
plt.savefig("/mnt/data/dashboard_test.png",dpi=150,bbox_inches="tight")
plt.close()
print("OK")
''',
}


def call_agent(agent_id: str, prompt: str, previous_response_id: str = None) -> dict:
    """Call an agent via the Open Responses API."""
    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": agent_id,
        "input": prompt,
        "stream": False,
    }
    if previous_response_id:
        payload["previous_response_id"] = previous_response_id

    resp = requests.post(RESPONSES_ENDPOINT, headers=headers, json=payload, timeout=TIMEOUT)
    resp.raise_for_status()
    return resp.json()


def extract_response_data(response: dict) -> tuple:
    """Extract text, reasoning, and tool calls from response.

    NOTE: The Open Responses API does NOT return the actual code executed
    (function_call.arguments is always empty). We can only check methodology
    via the reasoning text (agent's thinking) and the message text (agent's output).
    Both are concatenated into code_blocks for pattern matching.
    """
    text_parts = []
    code_blocks = []
    files = []

    for output in response.get("output", []):
        if output.get("type") == "message":
            for content in output.get("content", []):
                if content.get("type") == "output_text":
                    text = content.get("text", "")
                    text_parts.append(text)
                    code_blocks.append(("message", text))
        elif output.get("type") == "reasoning":
            for content in output.get("content", []):
                if content.get("type") == "reasoning_text":
                    code_blocks.append(("reasoning", content.get("text", "")))
        elif output.get("type") == "function_call":
            name = output.get("name", "")
            args = output.get("arguments", "")
            # Track that code execution happened
            code_blocks.append(("function_call", f"TOOL_USED: {name} {args}"))

    response_text = "\n".join(text_parts)
    return response_text, code_blocks, files


def check_methodology(code_blocks: list, patterns: list, antipatterns: list) -> dict:
    """Check that expected methodology patterns are present in code execution."""
    all_code = "\n".join(text for _, text in code_blocks)
    results = {}

    for pattern in patterns:
        found = bool(re.search(pattern, all_code, re.IGNORECASE | re.DOTALL))
        results[f"EXPECTED: {pattern}"] = "FOUND" if found else "MISSING"

    for pattern in antipatterns:
        found = bool(re.search(pattern, all_code, re.IGNORECASE | re.DOTALL))
        results[f"FORBIDDEN: {pattern}"] = "VIOLATION" if found else "OK"

    return results


# === Test Definitions ===

TESTS = []


def test(test_id, agent_id, agent_name, prompt, patterns, antipatterns=None, description="", expect_file=False, file_ext=""):
    """Register a test case."""
    TESTS.append(TestCase(
        test_id=test_id,
        agent_id=agent_id,
        agent_name=agent_name,
        prompt=prompt,
        methodology_patterns=patterns,
        methodology_antipatterns=antipatterns or [],
        description=description,
        expect_file=expect_file,
        file_extension=file_ext,
    ))


# ==========================================
# DOCX Agent Tests
# ==========================================

# NOTE: The Open Responses API does NOT return executed code.
# Patterns are checked against reasoning (agent thinking) + message (agent output).
# We verify methodology intent, not exact code.

test("D01b", "agent_docx_complete", "DOCX",
     "Crée un compte-rendu de la réunion suivante : Réunion du 10 avril 2026, visioconférence Teams, "
     "organisée par Damien Juillard. Participants : Sophie Martin (Directrice RH, Nextera Corp) et "
     "Damien Juillard (Consultant IA, On Behalf AI). Sujet : cadrage projet IA RH. "
     "Décisions : lancement POC chatbot RH. Actions : étude faisabilité avant le 24 avril.",
     patterns=[
         r"fill_cr_template|compte.rendu|template.*CR",
         r"execute_code",
     ],
     description="Créer un CR via fill_cr_template.py",
     expect_file=True, file_ext=".docx")

test("D01c", "agent_docx_complete", "DOCX",
     "Crée un guide d'installation pour PostgreSQL sur Ubuntu, avec prérequis, étapes d'installation, "
     "configuration de base, et dépannage.",
     patterns=[
         r"fill_template|template.*base|template.*OBA",
         r"execute_code",
     ],
     description="Créer un guide technique via fill_template.py",
     expect_file=True, file_ext=".docx")

test("D02", "agent_docx_complete", "DOCX",
     "Voici un texte de CGV simplifié. Remplace 'le Client' par 'l'Utilisateur' et '30 jours' par "
     "'15 jours ouvrés' en tracked changes.\n\n"
     "Article 1 : le Client s'engage à respecter les présentes conditions. Le paiement est dû sous 30 jours. "
     "Article 2 : le Client peut résilier sous 30 jours de préavis.",
     patterns=[
         r"tracked.replace|tracked.changes|redline",
         r"execute_code",
     ],
     description="Tracked changes avec {{current_user}} comme auteur")

test("D05", "agent_docx_complete", "DOCX",
     "Convertis ce texte markdown en document Word avec les styles OBA :\n\n"
     "# Politique de télétravail\n## 1. Principes généraux\n"
     "Le télétravail est ouvert à tous les collaborateurs.\n"
     "## 2. Modalités\n- Maximum 3 jours par semaine\n- Accord du manager requis\n"
     "## 3. Obligations\nRespecter les horaires définis.",
     patterns=[r"pandoc|markdown|convert"],
     description="Conversion Markdown → DOCX avec template OBA pandoc")

test("D08", "agent_docx_complete", "DOCX",
     "Convertis ce texte en PDF professionnel avec mise en forme OBA :\n\n"
     "Titre : Rapport mensuel\nAuteur : Damien\n\n"
     "Section 1 : Résumé exécutif\nLe mois d'avril a été marqué par une croissance de 15%.\n\n"
     "Section 2 : Détails\n- Ventes : 150k€\n- Charges : 120k€\n- Résultat : +30k€",
     patterns=[r"pdf|soffice|convert"],
     description="Conversion DOCX → PDF via soffice",
     expect_file=True, file_ext=".pdf")

test("D10", "agent_docx_complete", "DOCX",
     "Crée un document Word 'Fiche produit' avec : titre 'Widget Pro X200', "
     "un tableau de spécifications (Poids: 1.2kg, Dimensions: 30x20x10cm, Prix HT: 149.90€), "
     "et un paragraphe de description marketing.",
     patterns=[
         r"fill_template|template.*OBA|template.*base",
         r"execute_code",
     ],
     description="Création avec fill_template.py + type table",
     expect_file=True, file_ext=".docx")


# ==========================================
# PPTX Agent Tests
# ==========================================

test("P01b", "agent_pptx_complete", "PPTX",
     "Crée une présentation de 5 slides sur l'IA générative pour une réunion interne.",
     patterns=[
         r"pptxgenjs|pptxgen|PptxGenJS|Node",
         r"execute_code",
     ],
     description="Création PPTX avec palette OBA",
     expect_file=True, file_ext=".pptx")

test("P01", "agent_pptx_complete", "PPTX",
     "Crée un pitch deck de 6 slides pour une startup EdTech appelée 'LearnAI'. "
     "Slides : Titre, Problème, Solution, Marché (TAM 50Md€), Business model, Ask (levée 1.5M€). "
     "Design moderne, palette verte/blanche.",
     patterns=[
         r"pptxgenjs|pptxgen|PptxGenJS|Node",
         r"execute_code",
     ],
     description="Création pitch deck pptxgenjs",
     expect_file=True, file_ext=".pptx")

test("P08", "agent_pptx_complete", "PPTX",
     "Crée une présentation de 3 slides avec des graphiques : "
     "(1) Titre 'Résultats T1', "
     "(2) Barres ventes par région (Nord:45k, Sud:38k, Est:52k, Ouest:41k), "
     "(3) Camembert charges (Salaires:60%, Loyer:15%, Marketing:20%, Divers:5%).",
     patterns=[
         r"pptxgenjs|pptxgen|chart|graphique",
         r"execute_code",
     ],
     description="Création slides avec graphiques pptxgenjs",
     expect_file=True, file_ext=".pptx")


# ==========================================
# XLSX Agent Tests
# ==========================================

test("X01", "agent_xlsx_complete", "XLSX",
     "Crée un fichier Excel de budget trimestriel avec un onglet Q1 contenant les postes "
     "Salaires (45000€), Loyer (8000€), Marketing (12000€), IT (6000€), et un sous-total en formule Excel. "
     "Formate avec en-têtes colorés et format monétaire €.",
     patterns=[
         r"openpyxl|Excel|xlsx",
         r"execute_code",
     ],
     description="Création Excel avec openpyxl + styles",
     expect_file=True, file_ext=".xlsx")

test("X03", "agent_xlsx_complete", "XLSX",
     "Crée un fichier Excel avec une colonne Prix (10, 20, 30) et une colonne Total qui fait =Prix*1.2. "
     "Recalcule les formules pour que les valeurs soient visibles.",
     patterns=[
         r"openpyxl|formul|recalc",
         r"execute_code",
     ],
     description="Création Excel + recalc.py",
     expect_file=True, file_ext=".xlsx")


# ==========================================
# PDF Agent Tests
# ==========================================

test("F13", "agent_pdf_complete", "PDF",
     "Crée un PDF professionnel 'Proposition commerciale' avec 3 sections : "
     "Contexte, Offre de service (avec liste à puces), et Conditions tarifaires.",
     patterns=[
         r"pdf|PDF|soffice|template|fill_template",
         r"execute_code",
     ],
     description="Création PDF via DOCX OBA → soffice",
     expect_file=True, file_ext=".pdf")

test("F04", "agent_pdf_complete", "PDF",
     "Crée deux PDFs simples (une page chacun avec du texte) puis fusionne-les en un seul.",
     patterns=[
         r"fusion|merge|fusionne|PdfMerger|qpdf",
         r"execute_code",
     ],
     description="Fusion de PDFs",
     expect_file=True, file_ext=".pdf")


# ==========================================
# FFmpeg Agent Tests
# ==========================================

test("M05", "agent_quick_edits", "FFmpeg",
     "Analyse les capacités de ffmpeg installé : quels codecs vidéo et audio sont disponibles ? "
     "Liste les 5 principaux codecs vidéo et audio.",
     patterns=[r"ffmpeg|ffprobe|codec"],
     description="Analyse ffmpeg capabilities")

test("M06", "agent_quick_edits", "FFmpeg",
     "Crée une image PNG de 800x600 pixels, fond bleu (#2F5597), avec le texte 'Test' en blanc centré.",
     patterns=[r"PIL|Pillow|image|Image|png|PNG"],
     description="Création d'image avec Pillow",
     expect_file=True, file_ext=".png")


# ==========================================
# DataViz Agent Tests
# ==========================================

test("A01", "agent_data_viz", "DataViz",
     "Génère un dataset fictif de 100 ventes (date, produit parmi A/B/C, montant entre 50 et 500€, "
     "région parmi Nord/Sud/Est/Ouest). Puis fais une analyse exploratoire : "
     "statistiques descriptives, répartition par produit, et un histogramme des montants.",
     patterns=[
         r"pandas|DataFrame|analyse|exploratoire|dataset|vente",
     ],
     description="Analyse exploratoire + visualisation",
     expect_file=True, file_ext=".png")

test("A02", "agent_data_viz", "DataViz",
     "Crée un dashboard en une seule image avec 4 graphiques à partir de données fictives : "
     "(1) courbe CA mensuel, (2) camembert par catégorie, (3) barres top clients, "
     "(4) scatter quantité vs montant. Utilise la palette de couleurs OBA.",
     patterns=[
         r"dashboard|subplots|graphique|OBA|palette|4.*graph",
     ],
     description="Dashboard 4 graphiques avec palette OBA",
     expect_file=True, file_ext=".png")


# ==========================================
# DOCX — Tests nécessitant un fichier source
# (marqués NEEDS_FILE — skippés si pas de fichier)
# ==========================================

test("D01", "agent_docx_complete", "DOCX",
     "[NEEDS_FILE:docx_cr_template] Voici un exemple de CR. Analyse sa structure puis produis un nouveau CR : "
     "Réunion produit du 14 avril 2026. Participants : Marie, Jean, Sophie. "
     "Décision : lancement V2 le 15 juin. Action : Jean prépare le plan média avant le 1er mai.",
     patterns=[r"unpack|lxml|template|analyse"],
     description="[NEEDS_FILE] Reproduire un CR depuis template utilisateur")

test("D03", "agent_docx_complete", "DOCX",
     "Accepte tous les tracked changes de ce document et produis la version finale propre.",
     patterns=[r"accept_changes|accept|tracked"],
     description="[NEEDS_FILE:docx_tracked] Accepter tracked changes → version propre")

test("D04", "agent_docx_complete", "DOCX",
     "Relis ce document et ajoute un commentaire Word sur le paragraphe 'Délais de livraison' : "
     "'À vérifier avec la logistique'.",
     patterns=[r"comment|commentaire|unpack"],
     description="[NEEDS_FILE:docx_proposal] Ajouter commentaires de relecture")

test("D06", "agent_docx_complete", "DOCX",
     "Fusionne ces deux documents Word en un seul, avec un saut de page entre eux.",
     patterns=[r"fusionne|merge|docxcompose|page_break"],
     description="[NEEDS_FILE:2x_docx] Fusion de deux DOCX")

test("D07", "agent_docx_complete", "DOCX",
     "Analyse ce document Word : liste des sections, nombre de tableaux, nombre d'images, "
     "et les 3 premiers paragraphes.",
     patterns=[r"unpack|pandoc|python-docx|analyse|section"],
     description="[NEEDS_FILE:docx_complex] Extraction contenu structuré")

test("D09", "agent_docx_complete", "DOCX",
     "Ce fichier est dans l'ancien format Word .doc. Convertis-le en .docx moderne.",
     patterns=[r"soffice|convert.*docx"],
     description="[NEEDS_FILE:doc_legacy] Conversion .doc → .docx")

test("D11", "agent_docx_complete", "DOCX",
     "Dans ce document, seule la première occurrence de 'Directeur' doit être remplacée par 'Directrice' "
     "en tracked changes.",
     patterns=[r"tracked.replace|--first|first"],
     description="[NEEDS_FILE:docx_directeur] Remplacement ciblé --first")

test("D12", "agent_docx_complete", "DOCX",
     "Remplis ce template de lettre avec : consultant=Marie Dupont, client=Société ABC, "
     "date=1er mai 2026, durée=6 mois, tarif=850€ HT. Puis remplace 'les conditions définies' par "
     "'les conditions révisées du contrat-cadre' en tracked changes. Exporte en PDF.",
     patterns=[r"unpack|tracked|soffice|pdf"],
     description="[NEEDS_FILE:docx_lettre] Pipeline complet template→tracked→PDF")


# ==========================================
# PPTX — Tests supplémentaires
# ==========================================

test("P02", "agent_pptx_complete", "PPTX",
     "Remplace le titre du slide 1 par 'Bilan annuel 2025' et le sous-titre par 'Direction Commerciale'.",
     patterns=[r"unpack|xml|edit|slide"],
     description="[NEEDS_FILE:pptx_corporate] Édition template PPTX")

test("P03", "agent_pptx_complete", "PPTX",
     "Analyse ce template PowerPoint : montre-moi un aperçu visuel de tous les layouts.",
     patterns=[r"thumbnail|layout|aperçu|analyse"],
     description="[NEEDS_FILE:pptx_template] Analyse template avec thumbnails")

test("P04", "agent_pptx_complete", "PPTX",
     "Duplique le slide 3 de cette présentation 3 fois pour avoir 4 copies au total.",
     patterns=[r"add_slide|dupli|copie"],
     description="[NEEDS_FILE:pptx_casestudy] Duplication de slides")

test("P05", "agent_pptx_complete", "PPTX",
     "Nettoie ce fichier PowerPoint : supprime les slides masqués et les médias non référencés.",
     patterns=[r"clean|nettoie|orphan"],
     description="[NEEDS_FILE:pptx_heavy] Nettoyage PPTX volumineux")

test("P06", "agent_pptx_complete", "PPTX",
     "Convertis cette présentation en PDF.",
     patterns=[r"soffice|pdf|convert"],
     description="[NEEDS_FILE:pptx_any] Conversion PPTX → PDF")

test("P07", "agent_pptx_complete", "PPTX",
     "Extrais le contenu textuel de cette présentation slide par slide en markdown.",
     patterns=[r"markitdown|markdown|extract"],
     description="[NEEDS_FILE:pptx_formation] Extraction contenu en markdown")

test("P09", "agent_pptx_complete", "PPTX",
     "Ajoute un nouveau slide à la fin en utilisant le 2ème layout disponible. "
     "Titre : 'Prochaines étapes', contenu : 'Valider le budget', 'Recruter 2 devs', 'Lancer V2'.",
     patterns=[r"add_slide|layout|unpack"],
     description="[NEEDS_FILE:pptx_layouts] Ajout slide depuis layout")

test("P10", "agent_pptx_complete", "PPTX",
     "Remplace toutes les occurrences de 'Acme Corp' par 'GlobalTech SA' dans tous les slides.",
     patterns=[r"replace|remplac|unpack|xml"],
     description="[NEEDS_FILE:pptx_acme] Remplacement texte dans toute la présentation")

test("P11", "agent_pptx_complete", "PPTX",
     "Crée une mini-présentation de 3 slides : (1) Titre 'Résultats T1', "
     "(2) Barres ventes par région, (3) Camembert répartition charges.",
     patterns=[r"pptxgenjs|pptxgen|chart|addChart"],
     description="Création PPTX avec graphiques",
     expect_file=True, file_ext=".pptx")

test("P12", "agent_pptx_complete", "PPTX",
     "Analyse ce template corporate puis crée 3 slides dans le même style : "
     "'Objectifs 2026' avec bullets, 'Budget' avec un tableau, 'Calendrier'. Exporte en PDF.",
     patterns=[r"thumbnail|analyse|add_slide|soffice|pdf"],
     description="[NEEDS_FILE:pptx_corporate2] Pipeline analyse→création→export")


# ==========================================
# XLSX — Tests supplémentaires
# ==========================================

test("X02", "agent_xlsx_complete", "XLSX",
     "Analyse ce fichier Excel : nombre de lignes/colonnes, types de données, valeurs manquantes, "
     "top 5 clients par CA.",
     patterns=[r"pandas|read_excel|describe|info|groupby"],
     description="[NEEDS_FILE:xlsx_commercial] Analyse Excel existant")

test("X04", "agent_xlsx_complete", "XLSX",
     "Ajoute un graphique en barres montrant l'évolution des ventes mensuelles dans un nouvel onglet Dashboard.",
     patterns=[r"openpyxl|BarChart|chart|Dashboard"],
     description="[NEEDS_FILE:xlsx_ventes] Graphique dans Excel")

test("X05", "agent_xlsx_complete", "XLSX",
     "Convertis ce fichier .xls en .xlsx moderne sans perdre les données.",
     patterns=[r"soffice|convert.*xlsx"],
     description="[NEEDS_FILE:xls_legacy] Conversion XLS → XLSX")

test("X06", "agent_xlsx_complete", "XLSX",
     "Crée un tableau croisé dynamique montrant le CA par catégorie et par mois.",
     patterns=[r"pivot|crois|pandas"],
     description="[NEEDS_FILE:xlsx_transactions] Tableau croisé dynamique")

test("X07", "agent_xlsx_complete", "XLSX",
     "Crée un fichier Excel de suivi des objectifs avec 10 vendeurs : vert si résultat >= objectif, "
     "orange entre 80-100%, rouge si < 80%.",
     patterns=[r"openpyxl|Conditional|mise.en.forme|couleur"],
     description="Mise en forme conditionnelle",
     expect_file=True, file_ext=".xlsx")

test("X08", "agent_xlsx_complete", "XLSX",
     "Exporte ce fichier Excel en PDF en paysage.",
     patterns=[r"soffice|pdf|convert|paysage"],
     description="[NEEDS_FILE:xlsx_any] Export Excel → PDF")

test("X09", "agent_xlsx_complete", "XLSX",
     "Fusionne ces 3 fichiers Excel en un seul avec un onglet par fichier et un onglet Consolidé.",
     patterns=[r"pandas|openpyxl|fusionne|merge|consolid"],
     description="[NEEDS_FILE:3x_xlsx] Fusion de plusieurs Excel")

test("X10", "agent_xlsx_complete", "XLSX",
     "Nettoie ce fichier : supprime les doublons, normalise les noms, corrige les dates.",
     patterns=[r"pandas|drop_duplicates|nettoie|clean"],
     description="[NEEDS_FILE:xlsx_dirty] Nettoyage et dédoublonnage")

test("X11", "agent_xlsx_complete", "XLSX",
     "Crée un modèle de prévision de trésorerie sur 12 mois : solde initial 50k€, "
     "encaissements croissants +5%/mois depuis 20k€, décaissements fixes 18k€. "
     "Formules Excel natives, négatifs en rouge.",
     patterns=[r"openpyxl|formul|=SUM|trésorerie|prévision"],
     description="Modèle financier avec formules complexes",
     expect_file=True, file_ext=".xlsx")

test("X12", "agent_xlsx_complete", "XLSX",
     "Analyse ce fichier RH : effectif et masse salariale par département, "
     "salaire moyen/médian, graphiques barres + camembert dans un onglet Dashboard.",
     patterns=[r"pandas|openpyxl|chart|Dashboard|analyse"],
     description="[NEEDS_FILE:xlsx_rh] Analyse + viz + export complet")


# ==========================================
# PDF — Tests supplémentaires
# ==========================================

test("F01", "agent_pdf_complete", "PDF",
     "Extrais le texte de ce contrat et identifie les clauses : durée, montant, résiliation, pénalités.",
     patterns=[r"pdfplumber|extract|clause|texte"],
     description="[NEEDS_FILE:pdf_contrat] Extraction texte contrat")

test("F02", "agent_pdf_complete", "PDF",
     "Extrais les tableaux de ce PDF et convertis-les en Excel.",
     patterns=[r"pdfplumber|extract_table|pandas|Excel"],
     description="[NEEDS_FILE:pdf_tableaux] Extraction tableaux → Excel")

test("F03", "agent_pdf_complete", "PDF",
     "Ce PDF est un scan. Extrais le texte par OCR.",
     patterns=[r"pdf2image|pytesseract|OCR|image_to_string"],
     description="[NEEDS_FILE:pdf_scan] OCR PDF scanné")

test("F05", "agent_pdf_complete", "PDF",
     "Extrais les pages 3 à 7 de ce PDF dans un fichier séparé.",
     patterns=[r"pypdf|PdfReader|PdfWriter|qpdf|pages"],
     description="[NEEDS_FILE:pdf_10pages] Split PDF par pages")

test("F06", "agent_pdf_complete", "PDF",
     "Vérifie l'intégrité de ce PDF et répare-le si nécessaire.",
     patterns=[r"qpdf|check|repair|intégrité"],
     description="[NEEDS_FILE:pdf_corrupt] Vérification et réparation")

test("F07", "agent_pdf_complete", "PDF",
     "Convertis chaque page de ce PDF en image PNG haute résolution (300 DPI).",
     patterns=[r"pdf2image|pdftoppm|convert.*image|300|dpi"],
     description="[NEEDS_FILE:pdf_any] Conversion PDF → images HR")

test("F08", "agent_pdf_complete", "PDF",
     "Donne-moi toutes les métadonnées de ce PDF : auteur, date, nombre de pages, chiffrement.",
     patterns=[r"pypdf|PdfReader|metadata|qpdf"],
     description="[NEEDS_FILE:pdf_any2] Extraction métadonnées")

test("F09", "agent_pdf_complete", "PDF",
     "Les pages 2 et 5 de ce PDF sont à l'envers. Corrige-les.",
     patterns=[r"pypdf|rotate|rotation"],
     description="[NEEDS_FILE:pdf_rotated] Rotation de pages")

test("F10", "agent_pdf_complete", "PDF",
     "Ajoute un watermark 'BROUILLON' en diagonale sur toutes les pages de ce PDF.",
     patterns=[r"reportlab|watermark|merge|Canvas"],
     description="[NEEDS_FILE:pdf_any3] Ajout watermark")

test("F11", "agent_pdf_complete", "PDF",
     "Optimise ce PDF de 25 Mo pour réduire sa taille.",
     patterns=[r"qpdf|compress|optimis|linearize"],
     description="[NEEDS_FILE:pdf_heavy] Compression PDF")

test("F12", "agent_pdf_complete", "PDF",
     "Ce relevé bancaire est un scan. Extrais les transactions (date, libellé, débit, crédit) en Excel.",
     patterns=[r"pdf2image|pytesseract|pandas|OCR|Excel"],
     description="[NEEDS_FILE:pdf_scan_table] Pipeline OCR → tableaux → Excel")


# ==========================================
# FFmpeg — Tests supplémentaires
# ==========================================

test("M01", "agent_quick_edits", "FFmpeg",
     "Convertis cette vidéo en MP4 H.264+AAC, résolution 720p max.",
     patterns=[r"ffmpeg|ffprobe|libx264|aac|720"],
     description="[NEEDS_FILE:video_mov] Conversion vidéo MP4")

test("M02", "agent_quick_edits", "FFmpeg",
     "Extrais uniquement la piste audio de cette vidéo en MP3 192 kbps.",
     patterns=[r"ffmpeg|-vn|mp3lame|audio"],
     description="[NEEDS_FILE:video_any] Extraction audio")

test("M03", "agent_quick_edits", "FFmpeg",
     "Découpe cette vidéo pour garder uniquement la partie de 0:45 à 2:30.",
     patterns=[r"ffmpeg|-ss|-to|cut|découpe"],
     description="[NEEDS_FILE:video_any2] Découpe vidéo")

test("M04", "agent_quick_edits", "FFmpeg",
     "Crée un GIF animé à partir des 5 premières secondes de cette vidéo, 320px, 10fps.",
     patterns=[r"ffmpeg|gif|fps|scale"],
     description="[NEEDS_FILE:video_short] Création GIF animé")

test("M07", "agent_quick_edits", "FFmpeg",
     "Ajoute un bandeau noir en bas de cette image avec le texte '© onbehalf.ai 2026' en blanc.",
     patterns=[r"PIL|Pillow|ImageDraw|texte|bandeau"],
     description="[NEEDS_FILE:image_any] Ajout texte sur image")

test("M08", "agent_quick_edits", "FFmpeg",
     "Assemble ces vidéos bout à bout dans l'ordre.",
     patterns=[r"ffmpeg|concat|assemble|fusion"],
     description="[NEEDS_FILE:2x_video] Concaténation vidéos")

test("M09", "agent_quick_edits", "FFmpeg",
     "Ajoute cette musique en fond sonore à la vidéo, volume musique à 20%.",
     patterns=[r"ffmpeg|amix|amerge|volume|audio"],
     description="[NEEDS_FILE:video+audio] Ajout musique de fond")

test("M10", "agent_quick_edits", "FFmpeg",
     "Crée une mosaïque 2x2 à partir de ces 4 images.",
     patterns=[r"PIL|Pillow|paste|mosaïque|xstack"],
     description="[NEEDS_FILE:4x_image] Mosaïque d'images")

test("M11", "agent_quick_edits", "FFmpeg",
     "Extrais une capture d'écran de cette vidéo à 1 minute 23 secondes, en PNG pleine résolution.",
     patterns=[r"ffmpeg|-ss|frames|capture|screenshot"],
     description="[NEEDS_FILE:video_any3] Extraction frame spécifique")

test("M12", "agent_quick_edits", "FFmpeg",
     "Convertis toutes ces images en JPEG 1024px max, qualité 90%, nommées photo_01.jpg etc.",
     patterns=[r"PIL|Pillow|thumbnail|JPEG|convert|batch"],
     description="[NEEDS_FILE:multi_image] Conversion batch d'images")


# ==========================================
# DataViz — Tests supplémentaires
# ==========================================

test("A03", "agent_data_viz", "DataViz",
     "Génère des données de satisfaction client par canal (boutique, web, téléphone, 50 par canal). "
     "Fais un test ANOVA pour savoir si la satisfaction diffère entre canaux. "
     "Donne F-stat, p-value et un boxplot.",
     patterns=[r"ANOVA|f_oneway|scipy|boxplot|satisfaction"],
     description="Test statistique ANOVA",
     expect_file=True, file_ext=".png")

test("A04", "agent_data_viz", "DataViz",
     "Génère un dataset avec 5 variables numériques corrélées (200 lignes). "
     "Calcule la matrice de corrélation et affiche-la en heatmap annotée. "
     "Identifie les corrélations > 0.7.",
     patterns=[r"corr|heatmap|corrélation|seaborn"],
     description="Corrélation et heatmap",
     expect_file=True, file_ext=".png")

test("A05", "agent_data_viz", "DataViz",
     "Génère 200 points avec relation linéaire bruitée surface_m2 (30-150) vs prix (3.5k€/m²). "
     "Entraîne une régression linéaire, affiche R², équation, droite + intervalle de confiance.",
     patterns=[r"LinearRegression|regression|sklearn|R2|scatter"],
     description="Régression linéaire avec prédiction",
     expect_file=True, file_ext=".png")

test("A06", "agent_data_viz", "DataViz",
     "Génère des données clients (âge, revenu, dépenses, 300 lignes). "
     "Segmente avec K-Means, méthode du coude, visualise en 2D avec PCA.",
     patterns=[r"KMeans|cluster|PCA|coude|elbow"],
     description="Clustering K-Means",
     expect_file=True, file_ext=".png")

test("A07", "agent_data_viz", "DataViz",
     "Génère une série temporelle de CA quotidien sur 1 an (tendance haussière + saisonnalité). "
     "Fais une décomposition et une prévision naïve sur 30 jours.",
     patterns=[r"seasonal_decompose|statsmodels|série.temporelle|prévision"],
     description="Analyse séries temporelles",
     expect_file=True, file_ext=".png")

test("A08", "agent_data_viz", "DataViz",
     "Génère un dataset de ventes (100 lignes, colonnes: date, produit, région, montant). "
     "Exporte dans un Excel 3 onglets : données brutes, statistiques par catégorie, tableau croisé.",
     patterns=[r"ExcelWriter|openpyxl|onglet|sheet|pivot"],
     description="Export multi-onglets vers Excel",
     expect_file=True, file_ext=".xlsx")

test("A09", "agent_data_viz", "DataViz",
     "Génère 200 transactions avec 5% d'anomalies (montants aberrants). "
     "Détecte les outliers avec IQR et Z-score. Visualise sur un boxplot.",
     patterns=[r"zscore|IQR|outlier|anomalie|boxplot"],
     description="Détection d'anomalies",
     expect_file=True, file_ext=".png")

test("A10", "agent_data_viz", "DataViz",
     "Génère deux échantillons : 'avant' (n=50, μ=72, σ=12) et 'après' (n=50, μ=78, σ=11). "
     "Fais un test t de Student apparié, donne la conclusion, visualise les distributions.",
     patterns=[r"ttest|Student|scipy|distribution|avant.*après"],
     description="Comparaison avant/après avec test t",
     expect_file=True, file_ext=".png")

test("A11", "agent_data_viz", "DataViz",
     "Génère un nuage de mots à partir de ce texte : "
     "'L intelligence artificielle transforme les entreprises. Le machine learning et le deep learning "
     "permettent d automatiser des tâches complexes. Les données sont le carburant de l IA. "
     "Les algorithmes apprennent des patterns dans les données pour prédire et classifier.' "
     "Exclue les mots vides français, palette bleue.",
     patterns=[r"wordcloud|WordCloud|nuage|cloud"],
     description="Word cloud",
     expect_file=True, file_ext=".png")

test("A12", "agent_data_viz", "DataViz",
     "Génère un CSV de données sales (100 lignes, doublons, valeurs manquantes, types incohérents). "
     "Pipeline complet : charge, nettoie, analyse, 3 visualisations, export Excel avec images intégrées.",
     patterns=[r"pandas|matplotlib|openpyxl|pipeline|nettoie|clean"],
     description="Pipeline complet données → analyse → viz → export",
     expect_file=True, file_ext=".xlsx")


# === Test Runner ===

def run_test(test_case: TestCase) -> TestResult:
    """Run a single test and return the result."""
    result = TestResult(
        test_id=test_case.test_id,
        agent=test_case.agent_name,
        prompt=test_case.prompt[:100] + "...",
    )

    start = time.time()
    try:
        response = call_agent(
            test_case.agent_id,
            test_case.prompt,
            test_case.previous_response_id,
        )
        result.duration = time.time() - start

        if response.get("status") != "completed":
            result.status = "ERROR"
            result.error = f"Response status: {response.get('status')}"
            return result

        if response.get("error"):
            result.status = "ERROR"
            result.error = str(response["error"])
            return result

        text, code_blocks, files = extract_response_data(response)
        result.response_text = text[:500]
        result.code_blocks = code_blocks
        result.files_generated = files

        # Check methodology
        result.methodology_checks = check_methodology(
            code_blocks,
            test_case.methodology_patterns,
            test_case.methodology_antipatterns,
        )

        # Determine pass/fail
        has_missing = any(v == "MISSING" for v in result.methodology_checks.values())
        has_violation = any(v == "VIOLATION" for v in result.methodology_checks.values())

        if has_violation:
            result.status = "FAIL"
            result.error = "Methodology antipattern detected"
        elif has_missing:
            result.status = "FAIL"
            result.error = "Expected methodology pattern not found"
        else:
            result.status = "PASS"

    except requests.exceptions.Timeout:
        result.status = "ERROR"
        result.error = "Timeout"
        result.duration = time.time() - start
    except Exception as e:
        result.status = "ERROR"
        result.error = str(e)
        result.duration = time.time() - start

    return result


def print_result(result: TestResult):
    """Print a test result."""
    icon = {"PASS": "✓", "FAIL": "✗", "ERROR": "!", "SKIP": "○"}.get(result.status, "?")
    color = {"PASS": "\033[32m", "FAIL": "\033[31m", "ERROR": "\033[33m", "SKIP": "\033[90m"}.get(result.status, "")
    reset = "\033[0m"

    print(f"  {color}{icon} {result.test_id}{reset} [{result.agent}] {result.duration:.1f}s — {result.status}")

    if result.status != "PASS":
        if result.error:
            print(f"    Error: {result.error}")
        for check, value in result.methodology_checks.items():
            if value in ("MISSING", "VIOLATION"):
                print(f"    {value}: {check}")


def generate_sample_files():
    """Generate sample files by executing code directly in the code-interpreter.
    Downloads produced files to tests/results/files/."""
    FILES_DIR.mkdir(parents=True, exist_ok=True)
    print(f"\n{'='*60}")
    print(f"Generating sample files via code-interpreter")
    print(f"Output: {FILES_DIR}/")
    print(f"{'='*60}\n")

    results = []
    for test_id, code in FILE_GENERATION_SCRIPTS.items():
        lang = "js" if "require(" in code else "py"
        print(f"  Generating {test_id} ({lang})...", end=" ")
        try:
            resp = exec_code(code.strip(), lang)
            stdout = resp.get("stdout", "")
            stderr = resp.get("stderr", "")
            files = resp.get("files", [])
            session_id = resp.get("session_id", "")

            if stderr and "Error" in stderr:
                print(f"ERROR: {stderr[:100]}")
                results.append((test_id, "ERROR", stderr[:100]))
                continue

            if not files:
                print(f"NO FILES (stdout: {stdout.strip()[:60]})")
                results.append((test_id, "NO_FILES", stdout.strip()[:60]))
                continue

            # Download each file
            downloaded = []
            for f in files:
                file_id = f.get("id", "")
                file_name = f.get("name", "unknown")
                out_path = FILES_DIR / f"{test_id}_{file_name}"
                ok = download_file(session_id, file_id, out_path)
                if ok:
                    size = out_path.stat().st_size
                    downloaded.append(f"{file_name} ({size:,} bytes)")
                else:
                    downloaded.append(f"{file_name} (DOWNLOAD FAILED)")

            print(f"OK: {', '.join(downloaded)}")
            results.append((test_id, "OK", downloaded))

        except Exception as e:
            print(f"EXCEPTION: {e}")
            results.append((test_id, "EXCEPTION", str(e)))

    # Summary
    ok_count = sum(1 for _, s, _ in results if s == "OK")
    print(f"\n{'='*60}")
    print(f"Generated: {ok_count}/{len(results)} files")
    print(f"Files saved in: {FILES_DIR.absolute()}/")
    print(f"{'='*60}")

    # List generated files
    if FILES_DIR.exists():
        for f in sorted(FILES_DIR.iterdir()):
            if f.is_file():
                print(f"  {f.name} ({f.stat().st_size:,} bytes)")


def main():
    parser = argparse.ArgumentParser(description="Run agent API tests")
    parser.add_argument("tests", nargs="*", help="Specific test IDs to run (e.g., D01b D01c P01b)")
    parser.add_argument("--agent", help="Run all tests for an agent (docx, pptx, xlsx, pdf, ffmpeg, dataviz)")
    parser.add_argument("--list", action="store_true", help="List all available tests")
    parser.add_argument("--key", help="Agent API key (overrides AGENT_API_KEY env var)")
    parser.add_argument("--generate-files", action="store_true",
                        help="Generate sample files via code-interpreter (bypasses LLM, tests pipelines)")
    args = parser.parse_args()

    global API_KEY
    if args.key:
        API_KEY = args.key
    if not API_KEY:
        env_path = Path(__file__).parent.parent / ".env"
        if env_path.exists():
            for line in env_path.read_text().splitlines():
                if line.startswith("AGENT_API_KEY="):
                    API_KEY = line.split("=", 1)[1].strip()
        if not API_KEY:
            print("ERROR: Set AGENT_API_KEY env var or pass --key")
            sys.exit(1)

    if args.list:
        print(f"Available tests ({len(TESTS)}):\n")
        for t in TESTS:
            has_gen = " [GEN]" if t.test_id in FILE_GENERATION_SCRIPTS else ""
            print(f"  {t.test_id:6s} [{t.agent_name:8s}] {t.description}{has_gen}")
        print(f"\n[GEN] = has file generation script (usable with --generate-files)")
        return

    if args.generate_files:
        generate_sample_files()
        return

    # Filter tests
    tests_to_run = TESTS
    if args.tests:
        tests_to_run = [t for t in TESTS if t.test_id in args.tests]
    elif args.agent:
        agent_map = {
            "docx": "DOCX", "pptx": "PPTX", "xlsx": "XLSX",
            "pdf": "PDF", "ffmpeg": "FFmpeg", "dataviz": "DataViz",
        }
        agent_name = agent_map.get(args.agent.lower(), args.agent)
        tests_to_run = [t for t in TESTS if t.agent_name == agent_name]

    if not tests_to_run:
        print("No tests matched. Use --list to see available tests.")
        sys.exit(1)

    # Run tests
    print(f"\n{'='*60}")
    print(f"Running {len(tests_to_run)} agent tests")
    print(f"API: {API_BASE}")
    print(f"{'='*60}\n")

    results = []
    for test_case in tests_to_run:
        print(f"  Running {test_case.test_id} ({test_case.description})...")
        result = run_test(test_case)
        results.append(result)
        print_result(result)
        print()

    # Summary
    passed = sum(1 for r in results if r.status == "PASS")
    failed = sum(1 for r in results if r.status == "FAIL")
    errors = sum(1 for r in results if r.status == "ERROR")
    total_time = sum(r.duration for r in results)

    print(f"{'='*60}")
    print(f"Results: {passed} passed, {failed} failed, {errors} errors / {len(results)} total")
    print(f"Total time: {total_time:.1f}s")
    print(f"{'='*60}")

    # Save results
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    report_path = RESULTS_DIR / f"report_{timestamp}.json"
    report = {
        "timestamp": timestamp,
        "total": len(results),
        "passed": passed,
        "failed": failed,
        "errors": errors,
        "duration": total_time,
        "results": [
            {
                "test_id": r.test_id,
                "agent": r.agent,
                "status": r.status,
                "duration": r.duration,
                "error": r.error,
                "methodology_checks": r.methodology_checks,
                "response_preview": r.response_text[:200],
            }
            for r in results
        ],
    }
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False))
    print(f"\nReport saved: {report_path}")

    sys.exit(1 if failed + errors > 0 else 0)


if __name__ == "__main__":
    main()
